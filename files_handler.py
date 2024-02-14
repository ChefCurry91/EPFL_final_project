from flask import Flask, request, url_for, redirect
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font 
import os, datetime
import requests
from data_title_column_excel import dict_data_title_column


class FilesHandler():

    file_names=''
    excel_file=''

    #Initialize the path where the Excel file will be stored
    def __init__(self, path):
        self.path = path
        #self.excel_files = excel_files


        

    # For excel sheet manipulation, I relied on the following source:
    # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
    # https://automatetheboringstuff.com/2e/chapter13/

    def create_excel_file(self):

        wb = Workbook()

        # Get the active sheet (or specify a specific sheet if needed)
        sheet = wb.active

        # To customize font styles in cells,
        # import the Font() function from the openpyxl.styles module.
        
        # Create a font.
        bold_font_italic_16 = Font(size=16, bold=True, italic=True) 

        # Iterating through dictionnary
        # https://realpython.com/iterate-through-dictionary-python/
        

        # By iterating through the dictionary, we define the titles for each column 
        # and make adjustments to the font and size for each respective column.
        for letter,title in dict_data_title_column.items():
            sheet[letter+'1'] = title
            sheet[letter +'1'].font=bold_font_italic_16
            sheet.column_dimensions[letter].width =  35
            
       

        # Joining Paths with os.path.join
        # https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.
        file_path = os.path.join(self.path, self.excel_file)
        

        # Save operation applied to the Excel workbook
        wb.save(file_path)
        wb.close()

        # From here, we add the name of the first Excel sheet created in the Excel file
        # to the .txt file containing all the names corresponding to created Excel sheets
        file_containing_file_names = os.path.basename(self.file_names)


        # Get the full path to the file
        full_file_path = os.path.join(os.getcwd(), file_containing_file_names)

        # Check if a text file is not empty, 
        # Ensuring that no line break is added if the file contains no content
        # https://pythonhow.com/how/check-if-a-text-file-is-empty/#:~:text=getsize('nodata.,would%20output%20File%20is%20empty.
        
  
        if os.path.getsize(full_file_path) == 0:
            file_names = open(self.file_names,'a')
            file_names.write('Sheet.txt' + ' ')
            file_names.close()
        
        # Additionally, we create our first .txt file corresponding to the newly created Excel sheet
        # The first Excel sheet created is always named "Sheet"
        first_existing_text_file= 'Sheet.txt'
        my_new_file = open(first_existing_text_file, 'a')
        my_new_file.close()
        

        
        return file_path
    
    def create_excel_sheet(self,new_sheet_name):
        wb = Workbook()

        file_path = os.path.join(self.path, self.excel_file)


        # For excel sheet manipulation, I relied on the following source:
        # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
        # https://automatetheboringstuff.com/2e/chapter13/

        # Opening Excel Documents with OpenPyXL
        wb = load_workbook(file_path)


        #Define name of new sheet to be create
        title_new_sheet = new_sheet_name

        # Create new sheet
        new_sheet = wb.create_sheet(title=title_new_sheet)

        # To customize font styles in cells,
        # import the Font() function from the openpyxl.styles module.
        
        # Create a font.
        bold_font_italic_16 = Font(size=16, bold=True, italic=True) 

        # Iterating through dictionnary
        # https://realpython.com/iterate-through-dictionary-python/
        

        # By iterating through the dictionary, we define the titles for each column 
        # and make adjustments to the font and size for each respective column.
        for letter,title in dict_data_title_column.items():
            new_sheet[letter+'1'] = title
            new_sheet[letter +'1'].font=bold_font_italic_16
            new_sheet.column_dimensions[letter].width =  35
        
        
        wb.save(file_path)
        wb.close()
            

    def create_txt_file(self,sheet_name):

        # Create a new .txt file corresponding to a new Excel sheet created
        new_file_name= sheet_name + '.txt'
        my_new_file = open(new_file_name, 'a')
        my_new_file.close()
        
        # Add the names of the newly created text files to the "file_names" list
        # The goal is to store the names of text files corresponding to the Excel sheet in a single file

        file_names = open(self.file_names,'a')
        file_names.write(new_file_name + " ")
        file_names.close()


    def get_file_names(self):

        # Retrieve names stored in the "file_names.txt" file
        # The goal is to use these names for displaying options on the "posting-expenses" page.
        # This enables the user to choose the Excel sheet for posting expenses and to display any sheet on the "tracking-expenses" page.
        file_names=open(self.file_names)
        content = file_names.read()
        file_names.close()
        
        
        array_file_names = content.split(' ')

        # Remove white spaces from the array
        temp_array_file_names = array_file_names.pop()
        return array_file_names
    
    
    def get_payment_converted(self, currency_payment, amount, date_exchange_rate):
        
        # I relied on the documentation from the following web pages to use the Frankfurter API:

        # https://www.frankfurter.app/docs/#conversion
        # https://publicapis.io/frankfurter-api
        

        host = 'api.frankfurter.app'

        #url = f'https://{host}/{date_exchange_rate}?amount={amount}&from={currency_payment}&to={currency_conversion}'
        

        if currency_payment!='CHF':
            url = f'https://{host}/{date_exchange_rate}?amount={amount}&from={currency_payment}&to=CHF'

            # Send GET request to fetch conversion data:
            # https://www.geeksforgeeks.org/response-json-python-requests/
            response = requests.get(url)

            if response.status_code == 200:
                data = response.json()
                # Access elements of a Nested Dictionary:
                # https://www.programiz.com/python-programming/nested-dictionary
                converted_amount = data['rates']['CHF']
                
                # amount_converted is rounded to 2 decimal
                round(float(converted_amount), 2)
                return converted_amount
        else:
            return amount
        
    


    def add_data_to_excel_sheet(self,data_expense,excel_sheet_selected):


        bold_font_13 = Font(size=13, bold=True) 
        
        # Joining Paths with os.path.join
        # https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.
        
        # After testing, it was observed that when assigning 'file_name = self.excel_file', the application might crashe. 
        # JS functionanaliy preventFalseValues in "trackig-expenses" does not work. 
        # Therefore, to prevent this issue, the variable 'file_name' was initialized as 'file_name = 'expenses.xlsx'
        
        file_name = 'expenses.xlsx'
        #file_name =self.excel_file
        file_path = os.path.join(self.path,file_name)
        

        # For excel sheet manipulation, I relied on the following source:
        # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
        # https://automatetheboringstuff.com/2e/chapter13/

        # Opening Excel Documents with OpenPyXL
        wb = load_workbook(file_path)
        
        #Get the sheet where the data are going to be added
        
        active_sheet = wb[excel_sheet_selected]
        

        
        #sheet.max_row => get the maximum number of occupied rows 
        
        target_row = active_sheet.max_row + 1
        column_index= 1

        # Flag first row 
        skip_first_row = True 

        
        try:
            # Iterating through the dictionary 'data_expense',
            # which is initially populated upon clicking the "post" button
            # within the @app.route('/posting-expenses') endpoint

            # https://realpython.com/iterate-through-dictionary-python/

            for key,value in data_expense.items():
               
                # Considering that only cells from A1 to F1 are currently occupied,
                # and no cells in row 2 contain assigned values. As a result,
                # sheet.max_row is equal to 1, and the "target_row" variable is set to 2.

                # The variable "column_index" represents columns: 
                # "column_index" = 1 corresponds to the first column,
                # "column_index" = 2 corresponds to the second column on the spreadsheet.
                

                # While traversing the dictionary, we assign values to the next available row,
                # in this case, row 2. For each column from A to F (due to 6 key-value pairs),
                # the value stored in the variable "value" is assigned to the corresponding cell.
                # For example, the first key-value pair corresponds to cell A2, the second to B2, and so forth.


                active_sheet.cell(row=target_row, column=column_index).value = value
                active_sheet.cell(row=target_row, column=column_index).value = value
                
                # While the row remains unchanged, the column_index increments,
                # enabling movement to the next column while remaining on the same row
                column_index+=1


               # The subsequent code executes exclusively when the dictionary key corresponds to "amount_CHF".

                # The concept is that upon encountering the key "amount_CHF" in the last column, which is column F,
                # the script sums all the existing amounts present in column "E".
                if key == 'amount_CHF':

                    total_amount= 0

                    # While continuing through the dictionary, we've reached the final key-value pair,
                    # which corresponds to column "F". To aggregate all the amounts from column "E",
                    # we need to navigate or "move back" to column "E".
                   
                    column_E = column_index-1 

                    # Loop through every row in column "E"
                    # https://copyprogramming.com/howto/openpyxl-loop-through-rows-and-columns


                    for cell in active_sheet.iter_rows(min_col=column_E, max_col=column_E):
                        for row in cell:
                            
                            #Setting the variable "skip_first_row" to "True" ensures that during the initial iteration,
                            # we skip writing data to the first row of column "E", which contains the column title.

                            if skip_first_row:
                                skip_first_row=False


                            # While iterating through column "E" (representing different rows), 
                            # we retrieve the value assigned to each cell. 
                            # The retrieved value is then added to the cumulative total obtained from previous calculations  
                            else:
                                total_amount+=round(float(row.value), 2)
                                
                                # Subsequently, we assign the total amount to the cell in the next available row of column F.


                                active_sheet.cell(row=target_row, column=column_index).value = total_amount
                                active_sheet.cell(row=target_row, column=column_index).value = total_amount

                    
            wb.save(file_path)
            wb.close
            return total_amount
        
        except:
            return 
        
    def add_data_to_txt_file(self,data_expense,file_name):
        my_file = open(file_name, 'a')

        # To check if a file is empty using os.path.getsize, we need to provide the file name
        #https://www.geeksforgeeks.org/python-os-path-basename-method/
        
        my_file_name = os.path.basename(file_name)
       

        # Check if a text file is not empty, 
        # ensuring that no line break is added if the file contains no content
        # https://pythonhow.com/how/check-if-a-text-file-is-empty/#:~:text=getsize('nodata.,would%20output%20File%20is%20empty.
        
        if os.path.getsize(my_file_name) != 0:
            break_line ='\n'
            my_file.write(break_line)

        # content added to file
            
        # https://realpython.com/iterate-through-dictionary-python/

       # data_expense corresponds to the dictionary variable dict_expense where all the inputs entered in the form are stored.
       # We loop through it to extract only the values and add them to the .txt file corresponding to the Excel sheet.
       # The values are stored as strings for convenience.
            
            for key, value in data_expense.items():
               value_stringified= str(value) + ' '
               content_file = value_stringified
               my_file.write(content_file)
            my_file.close()

        
        else:
            for key, value in data_expense.items():
                value_stringified= str(value) + ' '
                content_file = value_stringified    
                my_file.write(content_file)
            my_file.close()

        return file_name


    def get_data_added_to_txt_file(self,txt_file_from_excel_sheet_selected):
        
        # Retrieve data from the .txt file corresponding to "file_excel_sheet_select"
        # The obtained data is returned as a two-dimensional array
        my_file = open(txt_file_from_excel_sheet_selected)
        content = my_file.read()
        my_file.close()
        array_expenses = content.split('\n')
        
    
        
        # We create and return the two-dimensional array as follows:
        
        two_dimensional_array_expenses = []
        for i in array_expenses:
            x = i.split()
            two_dimensional_array_expenses.append(x)
        return two_dimensional_array_expenses
    
    
    def clear_all_data(self,txt_file_from_excel_sheet_selected,excel_sheet_selected):
        my_file_to_delete = open(txt_file_from_excel_sheet_selected,'w')
        my_file_to_delete.close

        # Joining Paths with os.path.join
        # https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.
        file_path = os.path.join(self.path, self.excel_file)



        # For excel sheet manipulation, I relied on the following source:
        # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
        # https://automatetheboringstuff.com/2e/chapter13/

        # Opening Excel Documents with OpenPyXL
        wb = load_workbook(file_path)

        # Store the name of the selected sheet in the variable "active_sheet"
        active_sheet = wb[excel_sheet_selected]

        # If the selected sheet's name is "Sheet," delete all rows except the first one.
        # Ensuring that at least one sheet exists is the object
        active_sheet.delete_rows(2, active_sheet.max_row)


        wb.save(file_path)




    
    def delete_files(self,txt_file_from_excel_sheet_selected,excel_sheet_selected):

        # Joining Paths with os.path.join
        # https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.
        
        file_path = os.path.join(self.path, self.excel_file)
        print('hello')
        print(file_path)


        # For excel sheet manipulation, I relied on the following source:
        # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
        # https://automatetheboringstuff.com/2e/chapter13/

        
        # Opening Excel Documents with OpenPyXL
        wb = load_workbook(file_path)

        # Store the name of the selected sheet in the variable "active_sheet"
        active_sheet = wb[excel_sheet_selected]
        print(active_sheet)
        
        # If the selected sheet's name is "Sheet," delete all rows except the first one.
        # Ensuring that at least one sheet exists is the object

        if active_sheet.title == 'Sheet':
            return
       
            active_sheet.delete_rows(2, active_sheet.max_row)

            # Retrieve names stored in the "file_names.txt" file.
            # The objective is to create an array that serves as the basis for a new array,
            # excluding the name of the file to be deleted.
            file_names= open(self.file_names)
            content = file_names.read()
            file_names.close()


            array_file_names = content.split(' ')

            return array_file_name
            


        # If the selected sheet's name is not "Sheet," delete the corresponding text file and 
        # the sheet in the Excel file.

        elif active_sheet.title != 'Sheet':
            # delete the file
            print(txt_file_from_excel_sheet_selected)
            os.remove(txt_file_from_excel_sheet_selected)

            # Retrieve names stored in the "file_names.txt" file.
            # The objective is to create an array that serves as the basis for a new array,
            # excluding the name of the file to be deleted.
            file_names= open(self.file_names)
            content = file_names.read()
            file_names.close()
        
        
            array_file_names = content.split(' ')

            # Iterate through the array, appending only elements that do 
            # not correspond to the file to be deleted to a new array.

            temp = []
            for file_name in array_file_names:
                if file_name != txt_file_from_excel_sheet_selected:
                    temp.append(file_name)
            

            # Finally, the text file "file_names_file" is updated with the elements of the newly created array.
            file_names_file = open(self.file_names, 'w')
            file_names_file.write(' '.join(temp))
            file_names_file.close()

            return temp
        wb.save(file_path)
        
        

