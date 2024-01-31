from flask import Flask, request
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font 
import os, datetime
import requests
from data_title_column_excel import dict_data_title_column


class FileHandler():

    #Initialize the path where the Excel file will be stored
    def __init__(self, path,file_name,file_names):
        self.path = path
        self.file_name = file_name
        self.file_names = file_names


        

    # For excel sheet manipulation, I relied on the following source:
    # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
    # https://automatetheboringstuff.com/2e/chapter13/

    def create_excel_file(self):

        wb = Workbook()

        # Get the active sheet (or specify a specific sheet if needed)
        sheet = wb.active

        # To customize font styles in cells,
        # import the Font() function from the openpyxl.styles module.
        
        # Create a font.
        bold_font_italic_16 = Font(size=16, bold=True, italic=True) 

        # Iterating through dictionnary
        # https://realpython.com/iterate-through-dictionary-python/
        

        # By iterating through the dictionary, we define the titles for each column 
        # and make adjustments to the font and size for each respective column.
        for letter,title in dict_data_title_column.items():
            sheet[letter+'1'] = title
            sheet[letter +'1'].font=bold_font_italic_16
            sheet.column_dimensions[letter].width =  35
            
        file_name = 'expenses.xlsx'

        # Joining Paths with os.path.join
        # https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.
        file_path = os.path.join(self.path, self.file_name)

        # save operation applied to excell workbook
        wb.save(file_path)
        wb.close()

    
        file_containing_file_names = os.path.basename(self.file_names)

        # Get the full path to the file
        full_file_path = os.path.join(os.getcwd(), file_containing_file_names)
        print(full_file_path)

        # Check if a text file is not empty, 
        # ensuring that no line break is added if the file contains no content
        # https://pythonhow.com/how/check-if-a-text-file-is-empty/#:~:text=getsize('nodata.,would%20output%20File%20is%20empty.
        
  

        if os.path.getsize(full_file_path) == 0:
            file_names = open(self.file_names,'a')
            file_names.write('Sheet.txt' + ' ')
            file_names.close()
        
        return file_path
    
    def create_excel_sheet(self,new_sheet_name):
        wb = Workbook()

        file_name = 'expenses.xlsx'
        file_path = os.path.join(self.path, file_name)

        # For excel sheet manipulation, I relied on the following source:
        # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
        # https://automatetheboringstuff.com/2e/chapter13/

        # Opening Excel Documents with OpenPyXL
        wb = load_workbook(file_path)


        #Define name of new sheet to be create
        title_new_sheet = new_sheet_name

        # Create new sheet
        new_sheet = wb.create_sheet(title=title_new_sheet)

        # To customize font styles in cells,
        # import the Font() function from the openpyxl.styles module.
        
        # Create a font.
        bold_font_italic_16 = Font(size=16, bold=True, italic=True) 

        # Iterating through dictionnary
        # https://realpython.com/iterate-through-dictionary-python/
        

        # By iterating through the dictionary, we define the titles for each column 
        # and make adjustments to the font and size for each respective column.
        for letter,title in dict_data_title_column.items():
            new_sheet[letter+'1'] = title
            new_sheet[letter +'1'].font=bold_font_italic_16
            new_sheet.column_dimensions[letter].width =  35
        
        
        wb.save(file_path)
        wb.close()

        
        print('excel sheet created')
        #print(number_of_sheets)
    

    def create_txt_file(self,file_name):

        # Create a new txt file corresponding to a new excel sheet
        new_file_name= file_name + '.txt'
        my_new_file = open(new_file_name, 'a')
        my_new_file.close()
        
        # Add to "file_names", the names a the new file created
        # Aim is to store the names of the text files corresponding
        # to the excel sheet in one file

        file_names = open(self.file_names,'a')
        file_names.write(new_file_name + " ")
        file_names.close()


    def get_file_names_(self):
        file_names= open(self.file_names)
        content = file_names.read()
        file_names.close()
        
        
        array_file_names = content.split(' ')

        # remove white space from array
        temp_array_file_names = array_file_names.pop()
        return array_file_names
    
    
    def get_payment_converted(self, currency_payment, amount, date_exchange_rate):
        
        # I relied on the documentation from the following web pages to use the Frankfurter API:

        # https://www.frankfurter.app/docs/#conversion
        # https://publicapis.io/frankfurter-api
        

        host = 'api.frankfurter.app'

        #url = f'https://{host}/{date_exchange_rate}?amount={amount}&from={currency_payment}&to={currency_conversion}'
        

        if currency_payment!='CHF':
            url = f'https://{host}/{date_exchange_rate}?amount={amount}&from={currency_payment}&to=CHF'

            # Send GET request to fetch conversion data:
            # https://www.geeksforgeeks.org/response-json-python-requests/
            response = requests.get(url)

            if response.status_code == 200:
                data = response.json()
                # Access elements of a Nested Dictionary:
                # https://www.programiz.com/python-programming/nested-dictionary
                converted_amount = data['rates']['CHF']
                print(f"{amount} {currency_payment} = {converted_amount} CHF at {date_exchange_rate}")
                
                # amount_converted is rounded to 2 decimal
                round(float(converted_amount), 2)
                return converted_amount
        else:
            return amount
        
    


    def add_data_to_excel_sheet(self,data_expense,excel_sheet_selected):


        bold_font_13 = Font(size=13, bold=True) 
        
        # Joining Paths with os.path.join
        # https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.
        
        
        file_name = 'expenses.xlsx'
        file_path = os.path.join(self.path, file_name)
        

        # For excel sheet manipulation, I relied on the following source:
        # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
        # https://automatetheboringstuff.com/2e/chapter13/

        # Opening Excel Documents with OpenPyXL
        wb = load_workbook(file_path)
        
        #Get the sheet where the data are going to be added
        
        print(excel_sheet_selected)
        active_sheet = wb[excel_sheet_selected]
        
        #for sheet_name in wb.sheetnames:
         #   if sheet_name == excel_sheet_selected:
          #      active_sheet = wb[sheet_name]
                

        
        #sheet.max_row => get the maximum number of occupied rows 
        
        target_row = active_sheet.max_row + 1
        column_index= 1

        # Flag first row 
        skip_first_row = True 

        
        try:
            # Iterating through the dictionary 'data_expense',
            # which is initially populated upon clicking the "post" button
            # within the @app.route('/posting-expenses') endpoint

            # https://realpython.com/iterate-through-dictionary-python/

            for key,value in data_expense.items():
               
                # Considering that only cells from A1 to F1 are currently occupied,
                # and no cells in row 2 contain assigned values. As a result,
                # sheet.max_row is equal to 1, and the "target_row" variable is set to 2.

                # The variable "column_index" represents columns: 
                # "column_index" = 1 corresponds to the first column,
                # "column_index" = 2 corresponds to the second column on the spreadsheet.
                

                # While traversing the dictionary, we assign values to the next available row,
                # in this case, row 2. For each column from A to F (due to 6 key-value pairs),
                # the value stored in the variable "value" is assigned to the corresponding cell.
                # For example, the first key-value pair corresponds to cell A2, the second to B2, and so forth.


                active_sheet.cell(row=target_row, column=column_index).value = value
                active_sheet.cell(row=target_row, column=column_index).value = value
                
                # While the row remains unchanged, the column_index increments,
                # enabling movement to the next column while remaining on the same row
                column_index+=1


               # The subsequent code executes exclusively when the dictionary key corresponds to "amount_CHF".

                # The concept is that upon encountering the key "amount_CHF" in the last column, which is column F,
                # the script sums all the existing amounts present in column "E".
                if key == 'amount_CHF':

                    total_amount= 0

                    # While continuing through the dictionary, we've reached the final key-value pair,
                    # which corresponds to column "F". To aggregate all the amounts from column "E",
                    # we need to navigate or "move back" to column "E".
                   
                    column_E = column_index-1 

                    # Loop through every row in column "E"
                    # https://copyprogramming.com/howto/openpyxl-loop-through-rows-and-columns


                    for cell in active_sheet.iter_rows(min_col=column_E, max_col=column_E):
                        for row in cell:
                            
                            #Setting the variable "skip_first_row" to "True" ensures that during the initial iteration,
                            # we skip writing data to the first row of column "E", which contains the column title.

                            if skip_first_row:
                                skip_first_row=False


                            # While iterating through column "E" (representing different rows), 
                            # we retrieve the value assigned to each cell. 
                            # The retrieved value is then added to the cumulative total obtained from previous calculations  
                            else:
                                total_amount+=round(float(row.value), 2)
                                
                                # Subsequently, we assign the total amount to the cell in the next available row of column F.


                                active_sheet.cell(row=target_row, column=column_index).value = total_amount
                                active_sheet.cell(row=target_row, column=column_index).value = total_amount

                    
            wb.save(file_path)
            wb.close
            return total_amount
        
        except:
            return 
        
    def add_data_to_txt_file(self,data_expense,file_name):
        #my_file = open(self.file_name, 'a')
        my_file = open(file_name, 'a')

        # To check if a file is empty using os.path.getsize, we need to provide the file name
        #https://www.geeksforgeeks.org/python-os-path-basename-method/
        
        my_file_name = os.path.basename(file_name)
       

        # Check if a text file is not empty, 
        # ensuring that no line break is added if the file contains no content
        # https://pythonhow.com/how/check-if-a-text-file-is-empty/#:~:text=getsize('nodata.,would%20output%20File%20is%20empty.
        
        if os.path.getsize(my_file_name) != 0:
            print('Empty')
            break_line ='\n'
            my_file.write(break_line)

        # content added to file
            
        # https://realpython.com/iterate-through-dictionary-python/


            for key, value in data_expense.items():
               value_stringified= str(value) + ' '
               content_file = value_stringified
               my_file.write(content_file)
            my_file.close()

        
        else:
            for key, value in data_expense.items():
                value_stringified= str(value) + ' '
                content_file = value_stringified    
                my_file.write(content_file)
            my_file.close()

        return file_name


    def get_data_added_to_txt_file(self,file_excel_sheet_selected):
        #my_file= open(self.file_name)
        
        my_file = open(file_excel_sheet_selected)
        content = my_file.read()
        my_file.close()
        array_expenses = content.split('\n')
        
        #print(array_expenses)
        #return array_expenses
        
        
        
        two_dimensional_array_expenses = []
        for i in array_expenses:
            x = i.split()
            two_dimensional_array_expenses.append(x)
            print(two_dimensional_array_expenses)
        return two_dimensional_array_expenses
    
    
    def clear_all_data_txt_file(self):
        my_file_to_delete = open(self.file_name,'w')
        my_file_to_delete.close()

    
    def clear_all_data_excel_sheet(self):

        # Joining Paths with os.path.join
        # https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.
        
        file_name = 'expenses.xlsx'
        file_path = os.path.join(self.path, file_name)

        # For excel sheet manipulation, I relied on the following source:
        # Automate the Boring Stuff with Python: Practical Programming for Total Beginners by AI Sweighart
        # https://automatetheboringstuff.com/2e/chapter13/

        # Opening Excel Documents with OpenPyXL
        wb = load_workbook(file_path)
        
        sheet = wb['Sheet']

        sheet.delete_rows(1,sheet.max_row)
        wb.save(file_path)
        
        

