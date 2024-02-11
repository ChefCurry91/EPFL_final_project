from flask import Flask, request, session
import os
from openpyxl import Workbook,load_workbook
from files_handler import FilesHandler
from data_title_column_excel import dict_data_title_column
import datetime



# Joining Paths with os.path.join
# https://ioflood.com/blog/python-os-path/#:~:text=In%20this%20example%2C%20we're,to%20a%20user's%20documents%20directory.

# retrieving the user home directory
# https://www.tutorialspoint.com/How-to-find-the-real-user-home-directory-using-Python

desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
file = 'expenses_data.txt'
file_names = 'file_names.txt'
excel_file = 'expenses.xlsx'

my_files_handler= FilesHandler(desktop_path,excel_file,file,file_names)



my_files_handler.create_excel_file()




app = Flask('project-app')


def get_html(page_name):
    path= '/Users/fabio/Documents/TCC/s18-project-app/templates'
    complete_path = os.path.join(path, page_name)
    html_file = open(complete_path + '.html')
    content = html_file.read()
    html_file.close()
    return content


@app.route('/', methods = ['POST', 'GET'])
def index_page():
    page = get_html('index')

    sheet_name = request.form.get('sheet-name')
    if request.method == 'POST':
        my_files_handler.create_excel_sheet(sheet_name)
        my_files_handler.create_txt_file(sheet_name)


    return page

@app.route('/posting-expenses', methods = ['POST', 'GET'])
def post_expenses():
    page = get_html('posting-expenses')


    # Retrieve the names of files corresponding to sheets created in the Excel file
    # Each name is recorded in the text file "file_names.txt"
    file_names_array= my_files_handler.get_file_names_()
    
    # Create a <select> tag with all the sheet names contained in the Excel file

    actual_values='<div id="div-excel-sheet"><label>Excel Sheet</label> <select id = "list-excel-sheet" name="sheet-name"> <option> ---Choose Excel Sheet--- </option> ' 
    
    # Loop through the array file_names_array to retrieve all the sheet names created, 
    # and add each name as a value to the <option> tag

    for file_name in file_names_array:
        excel_sheet_name = file_name[:-4]
        actual_values+= f'<option value={excel_sheet_name}>{excel_sheet_name}</option>'

    actual_values+= '</select> </div>'
        



    # Retrieving HTML Form data using Flask
    # https://www.geeksforgeeks.org/retrieving-html-from-data-using-flask/

    
    amount = request.form.get('amount')
    expense= request.form.get('expense')
    date= request.form.get('date')
    currency_payment = request.form.get('currency-payment')
    excel_sheet_name_selected = request.form.get('sheet-name')
    

    

    if request.method == 'POST':
       
        
        # Add the suffix "txt" to the selected Excel sheet in the form
        # txt_file_name corresponds to a .txt file where the input entered 
        # in the form will be stored. The inputs were entered in the form located the "posting-expenses" page
        txt_file_name= excel_sheet_name_selected + '.txt'

        try:
            amount= round(float(amount), 2)
            converted_amount= my_files_handler.get_payment_converted(currency_payment,amount,date)
        
        except:
            amount=0
            converted_amount= my_files_handler.get_payment_converted(currency_payment,amount,date)
            

        # Parse the date string to datetime object
        # https://www.datacamp.com/tutorial/converting-strings-datetime-objects
        parsed_date = datetime.datetime.strptime(date, "%Y-%m-%d")

        # Format the date in the desired format (DD-MM-YYYY)
        # https://pynative.com/python-datetime-format-strftime/#h-how-to-format-date-and-time-in-python
        formatted_date = parsed_date.strftime("%d-%m-%Y")

        # Capitalize only the first letter
        expense= expense.capitalize()


        # Populating the dictonnary with input we get from the form.
        dict_expense = {
            'expense':expense,
            'date':formatted_date,
            'currency_payment':currency_payment,
            'amount':amount,
            'amount_CHF':converted_amount
        }

       
        my_files_handler.add_data_to_excel_sheet(dict_expense, excel_sheet_name_selected)
        my_files_handler.add_data_to_txt_file(dict_expense,txt_file_name)

       

    return page.replace('<p>Select Excel Sheet</p>',actual_values)



@app.route('/tracking-expenses', methods = ['POST', 'GET'])
def track_expenses():
    page= get_html('tracking-expenses')

    # Retrieve the names of files corresponding to sheets created in the Excel file
    # Each name is recorded in the text file "file_names.txt""
    file_names_array= my_files_handler.get_file_names_()


    # Create a <select> tag with all the sheet names contained in the Excel file

    current_values='<div id="div-excel-sheet-dropdown"><select id = "dropdown-excel-sheets-tracking-expenses" name="sheet-name-tracking"> <option> ---Select Your Excel Sheet--- </option>' 
    
    # Loop through the array file_names_array to retrieve all the sheet names created, 
    # and add each name as a value to the <option> tag

    for file_name in file_names_array:
        excel_sheet_name = file_name[:-4]
        current_values+= f'<option value="{excel_sheet_name}">{excel_sheet_name}</option>'

    current_values+= ' </select> <div id ="container-buttons"> <input id="submit-button-excel-sheet-tracking" type="submit" name = "input-sheet-tracking" value="Show Expenses"> <input id="submit-button-clear-inputs" type="submit" name="clear-all-inputs" value="Clear Data Sheet"> <input id="submit-button-delete-file" type="submit" name="delete_file" value="Delete Sheet"> </div></div>'

    # During the initialization of "file_excel_sheet_selected", we ensure that when landing on
    # the "tracking-expenses" page, the function get_data_added_to_text_file has an input
    # to prevent the page from crashing

    file_excel_sheet_selected = 'Sheet.txt'
    array_expenses = my_files_handler.get_data_added_to_txt_file(file_excel_sheet_selected)

    # Select the sheet we want to display for tracking
    excel_sheet_name_selected = request.form.get('sheet-name-tracking')
       
    track_expenses = request.form.get('input-sheet-tracking')
    clear_all_inputs = request.form.get('clear-all-inputs')
    delete_file = request.form.get('delete_file')


    if request.method == 'POST':

        if clear_all_inputs:
            txt_file_name=excel_sheet_name_selected + '.txt'
            my_files_handler.clear_all_data(txt_file_name,excel_sheet_name_selected)
        
        elif delete_file:
            txt_file_name=excel_sheet_name_selected + '.txt'

            if excel_sheet_name_selected != 'Sheet':
                new_file_names_array = my_files_handler.delete_files(txt_file_name,excel_sheet_name_selected)
            
                file_path = os.path.join(desktop_path,excel_file)

                # Opening Excel Documents with OpenPyXL
                wb = load_workbook(file_path)
                for file_name in new_file_names_array:
                    if file_name != txt_file_name:

                        # Remove the sheet only if it exists in the workbook
                        if excel_sheet_name_selected in wb.sheetnames:
                            sheet_to_delete = wb[excel_sheet_name_selected]
                            wb.remove(sheet_to_delete)

            # If the selected sheet's name is "Sheet," delete all rows except the first one.
            # Ensuring that at least one sheet exists in the object
                            
            elif excel_sheet_name_selected == 'Sheet':
                file_path = os.path.join(desktop_path,excel_file)
                wb = load_workbook(file_path)
                
                # Store the name of the selected sheet in the variable "active_sheet"
                active_sheet = wb[excel_sheet_name_selected]
                active_sheet.delete_rows(2, active_sheet.max_row)
                
                # Clear all the data in txt file corresponding to Excel sheet "Sheet"
                my_file_to_clear = open(txt_file_name,'w')
                my_file_to_clear.close


            wb.save(file_path)



        
        elif track_expenses:
       
            # Add the suffix "txt" to the selected Excel sheet in the form
            # txt_file_name corresponds to a .txt file where the input entered 
            # in the form will be stored. The inputs were entered in the form located the "posting-expenses" page
            txt_file_name= excel_sheet_name_selected + '.txt'

            # Retrieve data from the .txt file corresponding to "file_excel_sheet_select"
            # The obtained data is returned as a two-dimensional array
            array_expenses = my_files_handler.get_data_added_to_txt_file(txt_file_name)
    

    # From here, each individual expense added to the Excel file and retrieved
    # from the corresponding .txt file is displayed on the tracking-expense page
        
    actual_values=''

    # Label/Title associated with each respective piece of information

    
    label_list = ['Expense', 'Date', 'Currency','Payment', 'Pay. CHF' ]
    

    # Initialize variablew that will be incremented based on past user input

    aggregate_expense_in_CHF = 0
    total_amount_in_CHF= 0
    total_amount_in_EUR = 0
    total_amount_in_GBP = 0
    total_amount_in_USD=0
    
    # While looping through the two-dimensional array, we distinguish between
    # the first row and the subsequent ones for front-end purposes. We associate
    # labels to the first row.

    for count, inner_array in enumerate(array_expenses):

        actual_values+='<div class="container-per-expense">'
        
        # Count is used to distinguish the first row from the subsequent rows
        if count == 0:

            # Variable used as an index to loop through the label_list, changing the variable
            # accordingly based on the piece of data.

            index_label_list=0

            # Looping through first row to populate the front-end

            for element in inner_array:

                
                current_label =label_list[index_label_list]

                actual_values += f'<div id="container-label-title"><h4 id="label-{index_label_list}">{current_label}</h4><p id="value-{index_label_list}">{element}</p></div>'
                

                # Incrementing the previously initialized variable based on two conditions:
                # - The currency label
                # - The variable 'element' that represents the currency used for the payment

                if current_label == 'Pay. CHF':
                    aggregate_expense_in_CHF+= round(float(element),2)
                

                elif current_label == 'Currency' and element == 'CHF':
                    total_amount_in_CHF+= round(float(inner_array[3]),2)

                elif current_label == 'Currency' and element == 'EUR':
                    total_amount_in_EUR+= round(float(inner_array[3]),2)

                elif current_label == 'Currency' and element == 'GBP':
                    total_amount_in_GBP+= round(float(inner_array[3]),2)
                
                elif current_label == 'Currency' and element == 'USD':
                    total_amount_in_USD+= round(float(inner_array[3]),2)
                
                
                index_label_list+=1
            
            actual_values+='</div>'
        

        # Here, the logic applied to the first row is extended to each subsequent row

        else:
            index = 0
            
            for element in inner_array:
                
                current_label =label_list[index]

                actual_values += f'<div class="values"><span>{element}</span></div>'
                
                if current_label == 'Pay. CHF':
                    aggregate_expense_in_CHF+= round(float(element),2)

                elif current_label == 'Currency' and element == 'CHF':
                    total_amount_in_CHF+= round(float(inner_array[3]),2)
                
                elif current_label == 'Currency' and element == 'EUR':
                    total_amount_in_EUR+= round(float(inner_array[3]),2)
                
                elif current_label == 'Currency' and element == 'GBP':
                    total_amount_in_GBP+= round(float(inner_array[3]),2)
                
                elif current_label == 'Currency' and element == 'USD':
                    total_amount_in_USD+= round(float(inner_array[3]),2)
                
                index+= 1
                
            actual_values+='</div>'
    
    actual_values += '<div id="message-img-we-get-you-cover"><p id="get-you-cover-message">Like a Good Buddy, It Get You Covered</p><img id="img-tracking-expense-page" src ="./static/friendship.svg" alt="people-hanging-out"></div>'
    actual_values += '<div class ="container-aggregate-expenses">'
    actual_values += '<p id="total-payment-owerview">Total Payments Overview</p>'
    actual_values += f'<div class="aggregate-expenses-CHF"><label class="label-agreggate-expenses">Aggregate Expenses in CHF: </label> <p class="total-amount-expenses">{aggregate_expense_in_CHF}</p></div>'
    actual_values += f'<div class="aggregate-expenses"><label class="label-agreggate-expenses">Total Payments CHF: </label> <p class="total-amount-expenses"> {total_amount_in_CHF}</p></div>'
    actual_values += f'<div class="aggregate-expenses"><label class="label-agreggate-expenses">Total Payments €: </label> <p class="total-amount-expenses"> {total_amount_in_EUR}</p></div>'
    actual_values += f'<div class="aggregate-expenses"><label class="label-agreggate-expenses">Total Payments £:</label> <p class="total-amount-expenses"> {total_amount_in_GBP}</p></div>'
    actual_values += f'<div class="aggregate-expenses"><label class="label-agreggate-expenses">Total Payments $: </label> <p class="total-amount-expenses"> {total_amount_in_USD}</p></div>'
    actual_values += '</div>'

   
   

    
    return page.replace('<p>Track expenses</p>',actual_values).replace('<p>Select Excel Sheet</p>', current_values)


